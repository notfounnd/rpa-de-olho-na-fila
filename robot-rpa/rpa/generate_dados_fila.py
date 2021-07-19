import json
import requests
import openpyxl
from operator import itemgetter
from datetime import datetime

def get_date_time():
    global date_time

    date_time = datetime.now()
    date_time = date_time.strftime('%Y-%m-%d %H:%M:%S')

def get_dados_fila():
    headers = { 'Content-Type': 'application/x-www-form-urlencoded' }
    data = 'dados=dados'

    response = requests.post('https://deolhonafila.prefeitura.sp.gov.br/processadores/dados.php', headers=headers, data=data)
    json_payload = json.dumps(response.json(), indent=4, ensure_ascii=False)
    
    # print(json_payload)
    # print(response.json())

    return response.json()

def format_response_json(response):
    response = sorted(response, key=itemgetter('equipamento'))
    for element in response:
        print('************************')
        print(element['equipamento'])
        print(element['data_hora'])
        print(element['status_fila'])
    
    return response

def generate_cidades(path, response):
    sheet = 'CIDADES'

    workbook = openpyxl.load_workbook(filename=path)
    worksheet = workbook[sheet]
    
    for element in response:
        new_row = worksheet.max_row + 1
        worksheet.cell(column=1, row=int(new_row), value=str(element['equipamento']))
    
    workbook.save(path)

def get_local_info(local, response):
    local_info = ''

    for element in response:
        if local == element['equipamento']:
            if element['status_fila'] == 'SEM FILA':
                status_rank = 0
            elif element['status_fila'] == 'FILA PEQUENA':
                status_rank = 1
            elif element['status_fila'] == 'FILA MÉDIA':
                status_rank = 2
            elif element['status_fila'] == 'FILA GRANDE':
                status_rank = 3
            elif element['status_fila'] == 'AGUARDANDO ABASTECIMENTO':
                status_rank = 4
            elif element['status_fila'] == 'NÃO FUNCIONANDO':
                status_rank = 5
            else:
                status_rank = 9

            local_info = {
                'equipamento': element['equipamento'],
                'data_hora': element['data_hora'].split('.')[0],
                'status_fila': str(status_rank)
            }
    
    return local_info

def insert_status_fila(path, response):
    sheet_status = 'STATUS_FILA'

    workbook = openpyxl.load_workbook(filename=path)
    ws_status = workbook[sheet_status]

    column_status = ws_status.max_column + 1
    ws_status.cell(column=int(column_status), row=1, value=str(date_time))

    for row in ws_status.iter_rows(min_row=2):
        local = str(row[0].value)
        local_info = get_local_info(local, response)

        if local_info == '':
            local_info = { 'status_fila': 9 }

        ws_status.cell(column=int(column_status), row=row[0].row, value=str(local_info['status_fila']))
        # print(local + ' ---> ' + local_info['status_fila'])
        workbook.save(path)

    workbook.save(path)

def insert_status_time(path, response):
    sheet_time = 'STATUS_TIME'

    workbook = openpyxl.load_workbook(filename=path)
    ws_time = workbook[sheet_time]

    column_status = ws_time.max_column + 1
    ws_time.cell(column=int(column_status), row=1, value=str(date_time))

    for row in ws_time.iter_rows(min_row=2):
        local = str(row[0].value)
        local_info = get_local_info(local, response)

        if local_info == '':
            local_info = { 'data_hora': 'N/A' }
        
        ws_time.cell(column=int(column_status), row=row[0].row, value=str(local_info['data_hora']))
        # print(local + ' ---> ' + local_info['data_hora'])
        workbook.save(path)
        
    workbook.save(path)

def generate_data(path):
    get_date_time()
    response = get_dados_fila()
    response = format_response_json(response)
    insert_status_fila(path, response)
    insert_status_time(path, response)

# generate_data('./files/dados.xlsx')
