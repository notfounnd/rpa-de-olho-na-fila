import json
import openpyxl
import random

def generate_chart_labels(path):
    sheet = 'STATUS_FILA'

    workbook = openpyxl.load_workbook(filename=path)
    ws_status = workbook[sheet]

    labels = []
    for col in ws_status.iter_cols(min_col=2):
        label = col[0].value
        labels.append(label)
    
    workbook.close()

    with open('rpa/files/data_model.json') as json_file:
        json_object = json.load(json_file)
        json_file.close()
    
    json_object['labels'] = labels

    with open('rpa/files/data.json', 'w') as json_file:
        json.dump(json_object, json_file, indent=4)
        json_file.close()

def generate_chart_dataset(path):
    sheet = 'STATUS_FILA'

    workbook = openpyxl.load_workbook(filename=path)
    ws_status = workbook[sheet]

    for row in ws_status.iter_rows(min_row=2):
        color = '#' + str(generate_random_color())
        city = row[0].value
        data = []

        for col in ws_status.iter_cols(min_col=2):
            row_number = row[0].row - 1
            data.append(col[row_number].value)
        
        # print(city + " - " + str(data))
    
        with open('rpa/files/label_model.json') as json_file:
            label_object = json.load(json_file)
            json_file.close()

        label_object['label'] = city
        label_object['data'] = data
        label_object['backgroundColor'] = color
        label_object['borderColor'] = color

        with open('rpa/files/data.json') as json_file:
            json_data = json.load(json_file)
            json_file.close()

        json_data['dataset'].append(label_object)

        with open('rpa/files/data.json', 'w') as json_file:
            json.dump(json_data, json_file, indent=4)
            json_file.close()

    workbook.close()

def generate_random_color():
    color = '%06x' % random.randint(0, 0xFFFFFF)
    return color

def generate_chart(path):
    generate_chart_labels(path)
    generate_chart_dataset(path)

# generate_chart('./files/dados.xlsx')
